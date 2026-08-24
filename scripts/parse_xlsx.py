import json
import os
import re
import sys

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


HEADER_ALIASES = {
    "name": "full_name",
    "full name": "full_name",
    "address": "address",
    "add": "address",
    "nic": "nic",
    "nic number": "nic",
    "email": "email",
    "photo": "photo",
    "photo file path": "photo",
    "qr": "qr_code",
    "qr file path": "qr_code",
    "nickname": "nickname",
    "nick name": "nickname",
    "designation": "designation",
    "phone": "phone",
    "phone number": "phone",
    "issue date": "issue_date",
    "registration date": "issue_date",
    "expiry date": "expiry_date",
    "status": "status",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main(path):
    if not os.path.isfile(path):
        raise ValueError("The uploaded workbook could not be found.")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("The first worksheet is empty.")

    columns = {}
    for index, header in enumerate(headers):
        normalized = re.sub(r"\s+", " ", clean(header).lower())
        if normalized in HEADER_ALIASES:
            columns[HEADER_ALIASES[normalized]] = index

    required = ["full_name", "address", "nic", "email"]
    missing = [field for field in required if field not in columns]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))

    output = []
    for row_number, row in enumerate(rows, start=2):
        record = {field: clean(row[index] if index < len(row) else "") for field, index in columns.items()}
        if not any(record.values()):
            continue
        record["row_number"] = row_number
        record.setdefault("photo", "")
        record.setdefault("qr_code", "")
        record.setdefault("nickname", "")
        for field in ("designation", "phone", "issue_date", "expiry_date", "status"):
            record.setdefault(field, "")
        output.append(record)

    print(json.dumps({"sheet": sheet.title, "records": output}, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main(sys.argv[1])
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        sys.exit(1)
